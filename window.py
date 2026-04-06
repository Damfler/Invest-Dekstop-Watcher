"""
window.py — всплывающее окно по левому клику.

HTML-дашборд в стиле T-Bank Invest через pywebview (Edge WebView2).
HTML загружается из файла dashboard.html через url=file:// для полной
интерактивности (inline html= блокирует JS-события в EdgeChromium).
"""
import threading
import logging
import json
import os
from datetime import datetime

import webview as _webview

log = logging.getLogger("tbank.window")

# PyInstaller onefile: ресурсы в sys._MEIPASS, иначе рядом со скриптом
import sys as _sys
_BASE_DIR  = getattr(_sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
_HTML_FILE = os.path.join(_BASE_DIR, "dashboard.html")


# ──────────────────────────────────────────────────────────────────────────────
#  JS API для pywebview
# ──────────────────────────────────────────────────────────────────────────────
class _DashboardAPI:
    def __init__(self, store, refresh_callback, cfg):
        self._store   = store
        self._refresh = refresh_callback
        self._cfg     = cfg

    def get_data(self) -> str:
        s = self._store.snapshot()
        s["theme"] = self._cfg.get("theme", "system")
        s["use_logos"] = self._cfg.get("use_logos", False)
        s["bond_horizon_days"] = self._cfg.get("bond_horizon_days", 60)
        s["app_name"] = self._cfg.get("app_name", "")
        s["use_custom_icons"] = self._cfg.get("use_custom_icons", True)
        s["show_hints"] = self._cfg.get("show_hints", False)
        s["auto_update"] = self._cfg.get("auto_update", True)
        from version import APP_VERSION, APP_NAME
        s["app_version"] = APP_VERSION
        s["app_brand"]   = APP_NAME
        def _serial(obj):
            if isinstance(obj, datetime):
                return obj.isoformat()
            if isinstance(obj, set):
                return list(obj)
            raise TypeError(f"Not serializable: {type(obj)}")
        return json.dumps(s, default=_serial, ensure_ascii=False)

    def refresh(self):
        threading.Thread(target=self._refresh, daemon=True).start()

    def open_url(self, url: str):
        import webbrowser
        webbrowser.open(url)

    def export_csv(self):
        """Экспорт портфеля в CSV."""
        s = self._store.snapshot()
        positions = s.get("positions_extra", [])
        if not positions:
            return "no_data"
        import time as _t
        # Разделитель табуляция — Excel не путает с датами
        sep = "\t"
        lines = [sep.join(["Название", "Тикер", "ISIN", "Тип", "Кол-во", "Цена", "Стоимость", "P&L день", "P&L всё время", "Счёт"])]
        for p in positions:
            qty = p.get("qty", 0)
            qty_s = str(qty) if isinstance(qty, int) or qty == int(qty) else f"{qty:.2f}"
            lines.append(sep.join([
                str(p.get("name", "")),
                str(p.get("ticker", "")),
                str(p.get("isin", "")),
                str(p.get("instrumentType", "")),
                qty_s,
                f"{p.get('current_price', 0):.2f}".replace(".", ","),
                f"{p.get('current_value', 0):.2f}".replace(".", ","),
                f"{p.get('day_delta', 0):.2f}".replace(".", ","),
                f"{p.get('alltime_delta', 0):.2f}".replace(".", ","),
                str(p.get("account_name", "")),
            ]))
        fname = f"tbank_positions_{_t.strftime('%Y%m%d_%H%M%S')}.csv"
        if getattr(_sys, 'frozen', False):
            desktop = os.path.join(os.environ.get("USERPROFILE", ""), "Desktop")
        else:
            desktop = os.path.dirname(os.path.abspath(__file__))
        fpath = os.path.join(desktop, fname)
        with open(fpath, "w", encoding="utf-8-sig") as f:
            f.write("\n".join(lines))
        return fpath

    def export_xlsx(self):
        """Экспорт портфеля в настоящий Excel .xlsx с формулами."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, numbers
        import time as _t

        s = self._store.snapshot()
        positions = s.get("positions_extra", [])
        if not positions:
            return "no_data"

        wb = Workbook()

        # ── Лист "Позиции" ───────────────────────
        ws = wb.active
        ws.title = "Позиции"

        header_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        header_fill = PatternFill(start_color="1C1C1E", end_color="1C1C1E", fill_type="solid")
        money_fmt = '#,##0.00 "₽"'

        headers = ["Название", "Тикер", "ISIN", "Тип", "Кол-во", "Цена", "Стоимость", "P&L день", "P&L всё время", "Счёт"]
        widths = [30, 15, 15, 12, 10, 12, 15, 12, 15, 20]
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = w

        for r, p in enumerate(positions, 2):
            ws.cell(row=r, column=1, value=p.get("name", ""))
            ws.cell(row=r, column=2, value=p.get("ticker", ""))
            ws.cell(row=r, column=3, value=p.get("isin", ""))
            ws.cell(row=r, column=4, value=p.get("instrumentType", ""))
            ws.cell(row=r, column=5, value=p.get("qty", 0))
            c = ws.cell(row=r, column=6, value=p.get("current_price", 0))
            c.number_format = money_fmt
            # Стоимость = формула E*F
            c = ws.cell(row=r, column=7)
            c.value = f"=E{r}*F{r}"
            c.number_format = money_fmt
            c = ws.cell(row=r, column=8, value=p.get("day_delta", 0))
            c.number_format = money_fmt
            c = ws.cell(row=r, column=9, value=p.get("alltime_delta", 0))
            c.number_format = money_fmt
            ws.cell(row=r, column=10, value=p.get("account_name", ""))

        # Строка итого
        tr = len(positions) + 2
        ws.cell(row=tr, column=1, value="ИТОГО").font = Font(bold=True)
        for col in [5, 7, 8, 9]:
            letter = ws.cell(row=1, column=col).column_letter
            c = ws.cell(row=tr, column=col)
            c.value = f"=SUM({letter}2:{letter}{tr-1})"
            c.font = Font(bold=True)
            if col != 5:
                c.number_format = money_fmt

        # ── Лист "Счета" ─────────────────────────
        ws2 = wb.create_sheet("Счета")
        portfolios = s.get("portfolios", [])
        headers2 = ["Счёт", "Итого", "P&L день", "P&L всё время"]
        for i, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
        ws2.column_dimensions["A"].width = 25
        for col in "BCD":
            ws2.column_dimensions[col].width = 18

        for r, p in enumerate(portfolios, 2):
            ws2.cell(row=r, column=1, value=p.get("name", ""))
            c = ws2.cell(row=r, column=2, value=p.get("total", 0))
            c.number_format = money_fmt
            c = ws2.cell(row=r, column=3, value=p.get("day_delta", 0))
            c.number_format = money_fmt
            c = ws2.cell(row=r, column=4, value=p.get("alltime_delta", 0))
            c.number_format = money_fmt

        tr2 = len(portfolios) + 2
        ws2.cell(row=tr2, column=1, value="ИТОГО").font = Font(bold=True)
        for col_i, col_l in [(2, "B"), (3, "C"), (4, "D")]:
            c = ws2.cell(row=tr2, column=col_i)
            c.value = f"=SUM({col_l}2:{col_l}{tr2-1})"
            c.font = Font(bold=True)
            c.number_format = money_fmt

        # Сохранение
        fname = f"tbank_portfolio_{_t.strftime('%Y%m%d_%H%M%S')}.xlsx"
        if getattr(_sys, 'frozen', False):
            desktop = os.path.join(os.environ.get("USERPROFILE", ""), "Desktop")
        else:
            desktop = os.path.dirname(os.path.abspath(__file__))
        fpath = os.path.join(desktop, fname)
        wb.save(fpath)
        return fpath

    def export_excel(self):
        """Экспорт портфеля в Excel XML с формулами."""
        s = self._store.snapshot()
        positions = s.get("positions_extra", [])
        if not positions:
            return "no_data"

        xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
        xml += '<?mso-application progid="Excel.Sheet"?>\n'
        xml += '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"\n'
        xml += ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">\n'

        # Styles
        xml += '<Styles>\n'
        xml += '<Style ss:ID="Default"><Font ss:FontName="Segoe UI" ss:Size="10"/></Style>\n'
        xml += '<Style ss:ID="Header"><Font ss:FontName="Segoe UI" ss:Size="10" ss:Bold="1" ss:Color="#FFFFFF"/><Interior ss:Color="#1C1C1E" ss:Pattern="Solid"/></Style>\n'
        xml += '<Style ss:ID="Money"><NumberFormat ss:Format="#,##0.00"/></Style>\n'
        xml += '</Styles>\n'

        # Sheet: Positions
        xml += '<Worksheet ss:Name="Позиции">\n<Table>\n'
        # Column widths
        for w in [200, 80, 80, 60, 80, 100, 80, 80, 60, 120]:
            xml += f'<Column ss:Width="{w}"/>\n'

        # Header
        headers = ['Название', 'Тикер', 'ISIN', 'Тип', 'Кол-во', 'Цена', 'Стоимость', 'PnL день', 'PnL всё время', 'Счёт']
        xml += '<Row ss:StyleID="Header">\n'
        for h in headers:
            xml += f'<Cell><Data ss:Type="String">{h}</Data></Cell>\n'
        xml += '</Row>\n'

        # Data rows
        for i, p in enumerate(positions):
            row_num = i + 2  # 1-based, header is row 1
            xml += '<Row>\n'
            xml += f'<Cell><Data ss:Type="String">{_xml_esc(p.get("name", ""))}</Data></Cell>\n'
            xml += f'<Cell><Data ss:Type="String">{_xml_esc(p.get("ticker", ""))}</Data></Cell>\n'
            xml += f'<Cell><Data ss:Type="String">{_xml_esc(p.get("isin", ""))}</Data></Cell>\n'
            xml += f'<Cell><Data ss:Type="String">{_xml_esc(p.get("instrumentType", ""))}</Data></Cell>\n'
            xml += f'<Cell><Data ss:Type="Number">{p.get("qty", 0)}</Data></Cell>\n'
            xml += f'<Cell ss:StyleID="Money"><Data ss:Type="Number">{p.get("current_price", 0):.2f}</Data></Cell>\n'
            # Стоимость = формула: Кол-во × Цена (E × F)
            xml += f'<Cell ss:StyleID="Money" ss:Formula="=RC[-2]*RC[-1]"><Data ss:Type="Number">{p.get("current_value", 0):.2f}</Data></Cell>\n'
            xml += f'<Cell ss:StyleID="Money"><Data ss:Type="Number">{p.get("day_delta", 0):.2f}</Data></Cell>\n'
            xml += f'<Cell ss:StyleID="Money"><Data ss:Type="Number">{p.get("alltime_delta", 0):.2f}</Data></Cell>\n'
            xml += f'<Cell><Data ss:Type="String">{_xml_esc(p.get("account_name", ""))}</Data></Cell>\n'
            xml += '</Row>\n'

        # Totals row with formulas
        n = len(positions)
        xml += '<Row>\n'
        xml += '<Cell><Data ss:Type="String">ИТОГО</Data></Cell>\n'
        xml += '<Cell/><Cell/><Cell/>\n'
        # SUM qty
        xml += f'<Cell ss:Formula="=SUM(R2C:R{n+1}C)"><Data ss:Type="Number">0</Data></Cell>\n'
        xml += '<Cell/>\n'
        # SUM value
        xml += f'<Cell ss:StyleID="Money" ss:Formula="=SUM(R2C:R{n+1}C)"><Data ss:Type="Number">0</Data></Cell>\n'
        # SUM P&L day
        xml += f'<Cell ss:StyleID="Money" ss:Formula="=SUM(R2C:R{n+1}C)"><Data ss:Type="Number">0</Data></Cell>\n'
        # SUM P&L alltime
        xml += f'<Cell ss:StyleID="Money" ss:Formula="=SUM(R2C:R{n+1}C)"><Data ss:Type="Number">0</Data></Cell>\n'
        xml += '<Cell/>\n'
        xml += '</Row>\n'

        xml += '</Table>\n</Worksheet>\n'

        # Sheet: Summary
        portfolios = s.get("portfolios", [])
        xml += '<Worksheet ss:Name="Счета">\n<Table>\n'
        xml += '<Column ss:Width="200"/><Column ss:Width="120"/><Column ss:Width="120"/><Column ss:Width="120"/>\n'
        xml += '<Row ss:StyleID="Header">'
        for h in ['Счёт', 'Итого', 'PnL день', 'PnL всё время']:
            xml += f'<Cell><Data ss:Type="String">{h}</Data></Cell>'
        xml += '</Row>\n'
        for p in portfolios:
            xml += '<Row>'
            xml += f'<Cell><Data ss:Type="String">{_xml_esc(p.get("name", ""))}</Data></Cell>'
            xml += f'<Cell ss:StyleID="Money"><Data ss:Type="Number">{p.get("total", 0):.2f}</Data></Cell>'
            xml += f'<Cell ss:StyleID="Money"><Data ss:Type="Number">{p.get("day_delta", 0):.2f}</Data></Cell>'
            xml += f'<Cell ss:StyleID="Money"><Data ss:Type="Number">{p.get("alltime_delta", 0):.2f}</Data></Cell>'
            xml += '</Row>\n'
        # Totals
        n2 = len(portfolios)
        xml += '<Row><Cell><Data ss:Type="String">ИТОГО</Data></Cell>'
        for col in range(3):
            xml += f'<Cell ss:StyleID="Money" ss:Formula="=SUM(R2C:R{n2+1}C)"><Data ss:Type="Number">0</Data></Cell>'
        xml += '</Row>\n'
        xml += '</Table>\n</Worksheet>\n'

        xml += '</Workbook>'

        # Save
        import time as _t
        fname = f"tbank_portfolio_{_t.strftime('%Y%m%d_%H%M%S')}.xml"
        if getattr(_sys, 'frozen', False):
            # .exe — сохраняем на рабочий стол
            desktop = os.path.join(os.environ.get("USERPROFILE", ""), "Desktop")
        else:
            desktop = os.path.dirname(os.path.abspath(__file__))
        fpath = os.path.join(desktop, fname)

        with open(fpath, "w", encoding="utf-8") as f:
            f.write(xml)

        return fpath


    def set_theme(self, theme: str):
        from config import save_config
        self._cfg["theme"] = theme
        save_config(self._cfg)

    def set_use_logos(self, val: bool):
        from config import save_config
        self._cfg["use_logos"] = val
        save_config(self._cfg)

    def set_show_hints(self, val: bool):
        from config import save_config
        self._cfg["show_hints"] = val
        save_config(self._cfg)

    def set_custom_icons(self, val: bool):
        from config import save_config
        self._cfg["use_custom_icons"] = val
        save_config(self._cfg)

    def set_app_name(self, name: str):
        from config import save_config
        self._cfg["app_name"] = name
        save_config(self._cfg)

    def apply_update(self):
        """Скачать и применить обновление."""
        import os
        from updater import download_update, apply_update
        info = self._store.update_info
        if not info or not info.get("url"):
            return "no_update"
        exe_path = download_update(info["url"], info.get("asset_name", "InvestDesktopWatcher.exe"))
        if not exe_path:
            return "download_failed"
        apply_update(exe_path)
        os._exit(0)

    def set_auto_update(self, val: bool):
        from config import save_config
        self._cfg["auto_update"] = val
        save_config(self._cfg)

    def set_horizon(self, days: int):
        from config import save_config
        self._cfg["bond_horizon_days"] = days
        save_config(self._cfg)
        self._store.set_horizon(days, self._cfg)
        threading.Thread(target=self._refresh, daemon=True).start()

    # ── Управление подключениями ──────────────────────────────────────────────

    def get_connections(self) -> str:
        """Возвращает список подключений (без токенов) в JSON."""
        conns = self._cfg.get("connections", [])
        safe = [
            {
                "name":        c.get("name", ""),
                "broker":      c.get("broker", "tbank"),
                "enabled":     c.get("enabled", True),
                "use_sandbox": c.get("use_sandbox", False),
                "has_token":   bool(c.get("token")),
            }
            for c in conns
        ]
        return json.dumps(safe, ensure_ascii=False)

    def add_connection(self, name: str, broker: str, token: str,
                       use_sandbox: bool = False) -> bool:
        """Добавляет новое подключение. Требует перезапуска."""
        from config import save_config
        from constants import TOKEN_STUB
        if not token or token == TOKEN_STUB or len(token) < 10:
            return False
        conns = self._cfg.get("connections", [])
        conns.append({
            "name":        name.strip() or broker,
            "broker":      broker,
            "token":       token.strip(),
            "enabled":     True,
            "use_sandbox": bool(use_sandbox),
        })
        self._cfg["connections"] = conns
        save_config(self._cfg)
        return True

    def remove_connection(self, index: int) -> bool:
        """Удаляет подключение по индексу. Требует перезапуска."""
        from config import save_config
        conns = self._cfg.get("connections", [])
        if len(conns) <= 1:
            return False  # нельзя удалить последнее
        if 0 <= index < len(conns):
            conns.pop(index)
            self._cfg["connections"] = conns
            save_config(self._cfg)
            return True
        return False

    def toggle_connection(self, index: int) -> bool:
        """Включает/выключает подключение. Требует перезапуска."""
        from config import save_config
        conns = self._cfg.get("connections", [])
        if 0 <= index < len(conns):
            conns[index]["enabled"] = not conns[index].get("enabled", True)
            self._cfg["connections"] = conns
            save_config(self._cfg)
            return conns[index]["enabled"]
        return False


def _xml_esc(s):
    """Escape XML special chars."""
    if not s:
        return ""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


# ──────────────────────────────────────────────────────────────────────────────
#  Менеджер окна
# ──────────────────────────────────────────────────────────────────────────────
class DashboardWindow:
    """Дашборд через pywebview (Edge WebView2).

    HTML загружается из dashboard.html через url= для полной интерактивности.
    Окно создаётся скрытым при старте, toggle() показывает/скрывает.
    """

    def __init__(self, store, refresh_callback, cfg=None):
        self._store   = store
        self._refresh = refresh_callback
        self._visible = False
        self._cfg     = cfg or {}
        self._api     = _DashboardAPI(store, refresh_callback, self._cfg)

    def create_window(self):
        """Создать окно ДО webview.start(). Окно скрыто."""
        self._window = _webview.create_window(
            title            = "T-Bank Invest",
            url              = _HTML_FILE,
            js_api           = self._api,
            width            = 960,
            height           = 720,
            resizable        = True,
            easy_drag        = False,
            text_select      = True,
            hidden           = True,
            background_color = "#1c1c1e",
        )
        # Перехватываем закрытие: прячем окно вместо уничтожения
        self._window.events.closing += self._on_closing
        return self._window

    def _on_closing(self):
        """Кнопка X → прячем окно, не уничтожаем."""
        self._window.hide()
        self._visible = False
        return False  # False → closing.set() returns True → args.Cancel = True

    def toggle(self):
        """Показать/скрыть дашборд."""
        try:
            if self._visible:
                self._window.hide()
                self._visible = False
            else:
                self._visible = True
                self._window.show()
                try:
                    self._window.evaluate_js("if(typeof loadData==='function')loadData();")
                except Exception:
                    pass
        except Exception:
            log.exception("toggle failed")
            self._visible = False

    def request_quit(self):
        """Настоящее закрытие при выходе из приложения."""
        # Убираем обработчик closing чтобы destroy() не был отменён
        try:
            self._window.events.closing -= self._on_closing
        except Exception:
            pass
        try:
            self._window.destroy()
        except Exception:
            pass
